
import time
import pandas as pd
from openpyxl import load_workbook

FILE_PATH = r'c:/Users/Power/Desktop/datos_migracion/DATOS/2024/a0c2bff7-95d3-4d30-900a-4e7f231383fa-05122501533.xlsx'

def bench_pandas():
    start = time.time()
    try:
        xl = pd.ExcelFile(FILE_PATH)
        sheet_names = xl.sheet_names[:100] # Test 100 sheets
        print(f"Pandas: Scanning {len(sheet_names)} sheets...")
        for sheet in sheet_names:
            df = pd.read_excel(FILE_PATH, sheet_name=sheet, header=None)
    except Exception as e:
        print(e)
    end = time.time()
    print(f"Pandas Time: {end - start:.2f}s")

def bench_openpyxl():
    start = time.time()
    try:
        wb = load_workbook(FILE_PATH, read_only=True)
        sheet_names = wb.sheetnames[:100]
        print(f"OpenPyXL: Scanning {len(sheet_names)} sheets...")
        for sheet in sheet_names:
            ws = wb[sheet]
            # Just access first few rows
            rows = []
            for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
                 rows.append(row)
    except Exception as e:
        print(e)
    end = time.time()
    print(f"OpenPyXL Time: {end - start:.2f}s")

if __name__ == '__main__':
    bench_openpyxl()
    # bench_pandas() # We know this is slow
