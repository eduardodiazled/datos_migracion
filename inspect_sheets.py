
from openpyxl import load_workbook
import os

FILE_PATH = r'c:/Users/Power/Desktop/datos_migracion/DATOS/2021/a0c2bff7-95d3-4d30-900a-4e7f231383fa-041225222810.xlsx'

def list_sheets():
    try:
        wb = load_workbook(FILE_PATH, read_only=True)
        print(f"File: {os.path.basename(FILE_PATH)}")
        print(f"Total Sheets: {len(wb.sheetnames)}")
        print("First 50 Sheet Names:")
        for name in wb.sheetnames[:50]:
            print(f" - {name}")
        
        # Check for keywords like "Resumen", "Total", "Mes"
        summary_sheets = [s for s in wb.sheetnames if "resumen" in s.lower() or "total" in s.lower()]
        print(f"\nPotential Summary Sheets ({len(summary_sheets)}):")
        for s in summary_sheets:
            print(f" - {s}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    list_sheets()
