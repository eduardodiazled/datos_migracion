
import os
import re
import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
import unidecode

# Configuration
DATA_DIR = r'C:\Users\Power\Desktop\datos_migracion'
OUTPUT_FILE = 'transactions_dump.json'

def normalize_name(name):
    if not name: return ""
    name = re.sub(r'\s*\(.*?\)', '', str(name))
    try:
        norm = unidecode.unidecode(name).lower().strip()
    except:
        norm = str(name).lower().strip()
    return norm

def detect_service(desc):
    desc = str(desc).lower()
    if 'netflix' in desc or 'nfx' in desc: return 'NETFLIX'
    if 'disney' in desc: return 'DISNEY+'
    if 'prime' in desc: return 'PRIME VIDEO'
    if 'hbo' in desc or 'max' in desc: return 'HBO MAX'
    if 'plex' in desc: return 'PLEX'
    if 'iptv' in desc or 'magis' in desc: return 'IPTV'
    if 'combo' in desc: return 'COMBO'
    if 'spotify' in desc: return 'SPOTIFY'
    if 'youtube' in desc: return 'YOUTUBE'
    if 'start' in desc or 'star' in desc: return 'STAR+'
    if 'paramount' in desc: return 'PARAMOUNT+'
    if 'crunchyroll' in desc: return 'CRUNCHYROLL'
    return 'GENERICO'

def parse_treinta_excel(file_path, year):
    transactions = []
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        print(f"   -> Found {len(sheet_names)} sheets in {os.path.basename(file_path)}")
        
        last_valid_col_map = None
        
        for idx, sheet_name in enumerate(sheet_names):
             if idx % 200 == 0:
                print(f"      Scanning sheet {idx}/{len(sheet_names)}...", end='\r', flush=True)

             try:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                
                rows_buffer = []
                header_row_idx = None
                
                # Buffer first few rows to sniff header
                for i in range(20):
                    try:
                        row = next(rows_iter)
                        rows_buffer.append(row)
                        row_str = " ".join([str(x) for x in row if x is not None])
                        if "Fecha" in row_str and "Tipo" in row_str:
                            header_row_idx = i
                            break
                    except StopIteration:
                        break
                
                col_map = {}
                data_rows_source = [] 
                
                if header_row_idx is not None:
                     header = rows_buffer[header_row_idx]
                     for c_idx, val in enumerate(header):
                         val_str = str(val).lower() if val else ""
                         if "fecha" in val_str: col_map['date'] = c_idx
                         elif "contacto" in val_str or "cliente" in val_str: col_map['client'] = c_idx
                         elif "valor" in val_str or "monto" in val_str: col_map['amount'] = c_idx
                         elif "tipo" in val_str: col_map['type'] = c_idx
                         elif "descripci" in val_str: col_map['desc'] = c_idx
                     
                     if 'date' in col_map and 'amount' in col_map:
                         last_valid_col_map = col_map
                         data_rows_source = rows_buffer[header_row_idx+1:]
                          
                elif last_valid_col_map is not None:
                     if rows_buffer:
                         col_map = last_valid_col_map
                         data_rows_source = rows_buffer
                
                if 'date' in col_map and 'amount' in col_map:
                     
                     def chain_rows():
                         for r in data_rows_source: yield r
                         for r in rows_iter: yield r
                     
                     for row in chain_rows():
                         try:
                             date_val = None
                             amount = 0
                             tx_type = "Venta"
                             client = "Unknown"
                             desc = "GENERICO"
                             
                             valid_standard = False
                             try:
                                 if 'date' in col_map and 'amount' in col_map:
                                     date_raw = row[col_map['date']]
                                     amount_raw = row[col_map['amount']]
                                     
                                     if isinstance(date_raw, (datetime, pd.Timestamp)):
                                         date_val = date_raw
                                         valid_standard = True
                                     elif isinstance(date_raw, str):
                                          try:
                                              date_val = pd.to_datetime(date_raw, dayfirst=True)
                                              valid_standard = True
                                          except: pass
                                     
                                     if valid_standard:
                                         try:
                                             amount = float(amount_raw)
                                             client = row[col_map.get('client', 2)]
                                             tx_type = row[col_map.get('type', -1)] if 'type' in col_map else "Venta"
                                             desc = row[col_map.get('desc', -1)] if 'desc' in col_map else ""
                                             
                                             row_str_chk = " ".join([str(x) for x in row if x is not None]).lower()
                                             if "gasto" in row_str_chk or "egreso" in row_str_chk or "compra" in row_str_chk:
                                                 tx_type = "Gasto"
                                             
                                             if "anulado" in row_str_chk or "anulada" in row_str_chk:
                                                 continue
                                         except:
                                             valid_standard = False
                             except:
                                 valid_standard = False

                             if not valid_standard:
                                 val_str_full = " ".join([str(x) for x in row if x is not None])
                                 regex_match = re.search(r'(?:(\d{2}/\d{2}/\d{4})|(\d{4}-\d{2}-\d{2}))\s+(?:(?:\S+)\s+)?(.*?)\s+(\d+(?:\.\d+)?)\s*$', val_str_full)
                                 
                                 if regex_match and amount == 0:
                                     d_str = regex_match.group(1) if regex_match.group(1) else regex_match.group(2)
                                     date_val = pd.to_datetime(d_str, dayfirst=(regex_match.group(1) is not None))
                                     client = regex_match.group(3).strip()
                                     amount = float(regex_match.group(4))
                                     tx_type = "Venta"
                                     desc = val_str_full
                                     
                                     if "gasto" in val_str_full.lower() or "egreso" in val_str_full.lower(): 
                                         tx_type = "Gasto"
                                     if "anulado" in val_str_full.lower(): continue

                             if pd.isna(date_val): continue
                             
                             if str(tx_type).lower() == 'gasto':
                                 amount = -abs(amount) # Explicitly negative for logic, but we'll flag it as expense
                             
                             if abs(amount) > 100000000: continue
                             if amount == 0: continue
                             
                             service_detected = detect_service(desc if desc else "")
                             
                             transactions.append({
                                'source': 'Treinta',
                                'year': year,
                                'date': date_val.isoformat(),
                                'price': abs(amount),
                                'type': 'EGRESO' if amount < 0 else 'INGRESO',
                                'client_name': str(client) if client else 'Unknown',
                                'service': service_detected,
                                'description': str(desc)
                            })
                         except Exception:
                             continue
             except Exception:
                 pass
        
        wb.close()
        return transactions
    except Exception as e:
        print(f"⚠️ Error reading Excel {file_path}: {e}")
        return []

def main():
    print("Starting Export to JSON...")
    all_data = []
    
    # Process only relevant years or all
    for root, dirs, files in os.walk(DATA_DIR):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):
                path = os.path.join(root, file)
                try:
                    # Prefer explicit 2024 checks or folder name
                    folder_year = os.path.basename(root)
                    if folder_year.isdigit() and int(folder_year) >= 2021:
                        year = int(folder_year)
                        print(f"Scanning {file} ({year})")
                        txs = parse_treinta_excel(path, year)
                        all_data.extend(txs)
                except Exception as e:
                    print(e)
    
    print(f"Dumped {len(all_data)} records to {OUTPUT_FILE}")
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

if __name__ == '__main__':
    main()
