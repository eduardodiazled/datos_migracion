
import os
import re
import pandas as pd
import sqlite3
from bs4 import BeautifulSoup
from datetime import datetime
from fuzzywuzzy import process, fuzz
from openpyxl import load_workbook
import unidecode

# Configuration
DATA_DIR = r'C:\Users\Power\Desktop\datos_migracion'
DB_PATH = 'prisma/dev.db'

def normalize_name(name):
    if not name: return ""
    name = re.sub(r'\s*\(.*?\)', '', str(name))
    try:
        norm = unidecode.unidecode(name).lower().strip()
    except:
        norm = str(name).lower().strip()
    return norm

def detect_service(desc):
    desc = str(desc).lower()
    if 'netflix' in desc or 'nfx' in desc: return 'NETFLIX'
    if 'disney' in desc: return 'DISNEY+'
    if 'prime' in desc: return 'PRIME VIDEO'
    if 'hbo' in desc or 'max' in desc: return 'HBO MAX'
    if 'plex' in desc: return 'PLEX'
    if 'iptv' in desc or 'magis' in desc: return 'IPTV'
    if 'combo' in desc: return 'COMBO'
    if 'spotify' in desc: return 'SPOTIFY'
    if 'youtube' in desc: return 'YOUTUBE'
    if 'start' in desc or 'star' in desc: return 'STAR+'
    if 'paramount' in desc: return 'PARAMOUNT+'
    if 'crunchyroll' in desc: return 'CRUNCHYROLL'
    return 'GENERICO'

def parse_treinta_excel(file_path, year):
    transactions = []
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        print(f"   -> Found {len(sheet_names)} sheets in {os.path.basename(file_path)}")
        
        last_valid_col_map = None
        
        for idx, sheet_name in enumerate(sheet_names):
             if idx % 200 == 0:
                print(f"      Scanning sheet {idx}/{len(sheet_names)}...", end='\r', flush=True)

             try:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                
                rows_buffer = []
                header_row_idx = None
                
                # Buffer first few rows to sniff header
                for i in range(20):
                    try:
                        row = next(rows_iter)
                        rows_buffer.append(row)
                        row_str = " ".join([str(x) for x in row if x is not None])
                        if "Fecha" in row_str and "Tipo" in row_str:
                            header_row_idx = i
                            break
                    except StopIteration:
                        break
                
                col_map = {}
                data_rows_source = [] 
                
                if header_row_idx is not None:
                     # 1. FOUND HEADER -> New Map
                     header = rows_buffer[header_row_idx]
                     for c_idx, val in enumerate(header):
                         val_str = str(val).lower() if val else ""
                         if "fecha" in val_str: col_map['date'] = c_idx
                         elif "contacto" in val_str or "cliente" in val_str: col_map['client'] = c_idx # Prefer Contacto/Cliente
                         elif "valor" in val_str or "monto" in val_str: col_map['amount'] = c_idx
                         elif "tipo" in val_str: col_map['type'] = c_idx
                         elif "descripci" in val_str: col_map['desc'] = c_idx
                         # NOTE: explicitly ignoring "vendedor" to avoid capturing seller name
                     
                     if 'date' in col_map and 'amount' in col_map:
                         last_valid_col_map = col_map
                         data_rows_source = rows_buffer[header_row_idx+1:]
                         
                elif last_valid_col_map is not None:
                     # 2. NO HEADER -> Inherit Map
                     if rows_buffer:
                         col_map = last_valid_col_map
                         data_rows_source = rows_buffer
                
                if 'date' in col_map and 'amount' in col_map:
                     
                     def chain_rows():
                         for r in data_rows_source: yield r
                         for r in rows_iter: yield r
                     
                     for row in chain_rows():
                         try:
                             # 1. OPTIMISTIC: Standard Column Extraction
                             date_val = None
                             amount = 0
                             tx_type = "Venta"
                             client = "Unknown"
                             desc = "GENERICO"
                             
                             valid_standard = False
                             try:
                                 if 'date' in col_map and 'amount' in col_map:
                                     date_raw = row[col_map['date']]
                                     amount_raw = row[col_map['amount']]
                                     
                                     if isinstance(date_raw, (datetime, pd.Timestamp)):
                                         date_val = date_raw
                                         valid_standard = True
                                     elif isinstance(date_raw, str):
                                          try:
                                              date_val = pd.to_datetime(date_raw, dayfirst=True)
                                              valid_standard = True
                                          except: pass
                                     
                                     if valid_standard:
                                         try:
                                             amount = float(amount_raw)
                                             client = row[col_map.get('client', 2)]
                                             tx_type = row[col_map.get('type', -1)] if 'type' in col_map else "Venta"
                                             desc = row[col_map.get('desc', -1)] if 'desc' in col_map else ""
                                             
                                             row_str_chk = " ".join([str(x) for x in row if x is not None]).lower()
                                             if "gasto" in row_str_chk or "egreso" in row_str_chk or "compra" in row_str_chk:
                                                 tx_type = "Gasto"
                                             
                                             if "anulado" in row_str_chk or "anulada" in row_str_chk:
                                                 continue
                                         except:
                                             valid_standard = False
                             except:
                                 valid_standard = False

                             # 2. FALLBACK: Regex on JOINT ROW
                             val_str_full = " ".join([str(x) for x in row if x is not None])
                             
                             regex_match = re.search(r'(?:(\d{2}/\d{2}/\d{4})|(\d{4}-\d{2}-\d{2}))\s+(?:(?:\S+)\s+)?(.*?)\s+(\d+(?:\.\d+)?)\s*$', val_str_full)
                             
                             if regex_match and (not valid_standard or amount == 0):
                                 d_str = regex_match.group(1) if regex_match.group(1) else regex_match.group(2)
                                 date_val = pd.to_datetime(d_str, dayfirst=(regex_match.group(1) is not None))
                                 
                                 client = regex_match.group(3).strip()
                                 amount = float(regex_match.group(4))
                                 tx_type = "Venta"
                                 desc = val_str_full # Use full row as description for fallback
                                 
                                 if "gasto" in val_str_full.lower() or "egreso" in val_str_full.lower() or "compra" in val_str_full.lower(): 
                                     tx_type = "Gasto"
                                 
                                 if "anulado" in val_str_full.lower() or "anulada" in val_str_full.lower():
                                     continue

                                 valid_standard = True

                             if not valid_standard or pd.isna(date_val): continue
                             
                             # Handle Expenses
                             if str(tx_type).lower() == 'gasto':
                                 amount = -abs(amount)
                             
                             if abs(amount) > 100000000: continue
                             if amount == 0: continue
                             
                             # SERVICE DETECTION
                             service_detected = detect_service(desc if desc else val_str_full)
                             if service_detected == 'GENERICO':
                                 # Try detecting from full row if desc was empty
                                 service_detected = detect_service(val_str_full)

                             transactions.append({
                                'source': 'Treinta',
                                'year': year,
                                'date': date_val,
                                'price': amount,
                                'client_name': str(client) if client else 'Unknown',
                                'service': service_detected,
                                'is_renewal': True
                            })
                         except Exception:
                             continue
             except Exception:
                 pass
        
        wb.close()
        return transactions
    except Exception as e:
        print(f"⚠️ Error reading Excel {file_path}: {e}")
        return []

def migrate():
    print("Starting HIERARCHICAL Migration (Treinta ONLY)...")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    print("Clearing database...")
    cursor.execute("DELETE FROM 'Transaction'")
    cursor.execute("DELETE FROM Client")
    cursor.execute("DELETE FROM SalesProfile WHERE id >= 999")
    cursor.execute("DELETE FROM InventoryAccount WHERE id >= 999")
    conn.commit()

    service_map = {}
    next_account_id = 1000
    next_profile_id = 1000
    
    # TREINTA PASS
    print("\n--- PASS 1: PROCESSING TREINTA FILES (Primary Source) ---")
    for root, dirs, files in os.walk(DATA_DIR):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls'):
                path = os.path.join(root, file)
                try:
                    year = int(os.path.basename(root))
                except:
                    match = re.search(r'20\d{2}', file)
                    year = int(match.group(0)) if match else 2024
                
                print(f"Scanning Treinta: {file} ({year})")
                txs = parse_treinta_excel(path, year)
                
                for tx in txs:
                    service_name = tx['service']
                    
                    if service_name not in service_map:
                        now_ts = int(datetime.now().timestamp() * 1000)
                        cursor.execute('INSERT INTO InventoryAccount (id, servicio, tipo, email, password, createdAt, updatedAt) VALUES (?, ?, "ESTATICO", "migracion@estratosfera.net", "123", ?, ?)', (next_account_id, service_name, now_ts, now_ts))
                        cursor.execute('INSERT INTO SalesProfile (id, nombre_perfil, accountId, estado, createdAt, updatedAt) VALUES (?, ?, ?, "OCUPADO", ?, ?)', (next_profile_id, f"PERFIL_{service_name}", next_account_id, now_ts, now_ts))
                        service_map[service_name] = next_profile_id
                        next_account_id += 1
                        next_profile_id += 1
                    
                    profile_id = service_map[service_name]
                    client_name = tx['client_name']
                    norm_name = normalize_name(client_name)
                    dummy_phone = str(abs(hash(norm_name)))[:10]
                    
                    now_ts = int(datetime.now().timestamp() * 1000)
                    cursor.execute('INSERT OR IGNORE INTO Client (celular, nombre, createdAt, updatedAt) VALUES (?, ?, ?, ?)', (dummy_phone, client_name, now_ts, now_ts))
                    
                    start_date = tx['date']
                    if pd.isna(start_date): continue
                    
                    s_ts = int(start_date.timestamp() * 1000)
                    end_date = start_date + pd.Timedelta(days=30)
                    e_ts = int(end_date.timestamp() * 1000)
                    
                    cursor.execute('''
                        INSERT INTO "Transaction" (clienteId, perfilId, estado_pago, fecha_inicio, fecha_vencimiento, monto, createdAt, updatedAt)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (dummy_phone, profile_id, 'PAGADO', s_ts, e_ts, tx['price'], now_ts, now_ts))
                
                conn.commit()

    conn.close()
    print("Migration Complete!")

if __name__ == '__main__':
    migrate()
