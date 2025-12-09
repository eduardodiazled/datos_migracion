
import os
import pandas as pd
from openpyxl import load_workbook

def scan_2024_sheets():
    # Find 2024 file
    target_file = None
    for root, dirs, files in os.walk(r'c:\Users\Power\Desktop\datos_migracion'):
        for f in files:
            if '05122501533' in f:
                target_file = os.path.join(root, f)
                break
    
    if not target_file:
        print("2024 File not found")
        return

    print(f"Scanning file: {target_file}")
    
    try:
        wb = load_workbook(target_file, read_only=True)
        sheet_names = wb.sheetnames
        print(f"Total Sheets: {len(sheet_names)}")
        
        suspicious_sheets = []
        
        for name in sheet_names:
            name_lower = name.lower()
            if any(x in name_lower for x in ['gasto', 'egreso', 'compra', 'salida', 'resumen', 'total', 'balance']):
                suspicious_sheets.append(name)
        
        print(f"\nPotential Expense Sheets found by Name: {len(suspicious_sheets)}")
        for s in suspicious_sheets:
            print(f" - {s}")
            
        # If no named sheets, sample content of random sheets to see if they are separate types?
        # Actually, let's look at the first few sheets and last few sheets.
        
        print("\n--- Inspecting 'Suspicious' Sheets Content ---")
        if not suspicious_sheets:
            print("(None found. Listing first 10 sheets instead)")
            suspicious_sheets = sheet_names[:10]
            
        for s in suspicious_sheets[:5]: # limit to 5
            print(f"\nSheet: [{s}]")
            df = pd.read_excel(target_file, sheet_name=s, header=None, nrows=5)
            print(df.to_string())
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    scan_2024_sheets()
