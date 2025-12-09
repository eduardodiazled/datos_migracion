
from openpyxl import load_workbook
import os

FILE_PATH = r'c:/Users/Power/Desktop/datos_migracion/DATOS/2024/a0c2bff7-95d3-4d30-900a-4e7f231383fa-05122501533.xlsx'

def debug_dates():
    try:
        wb = load_workbook(FILE_PATH, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        print(f"Scanning 50 sheets for Date values...")
        
        date_samples = []
        
        for idx, sheet_name in enumerate(sheet_names[:50]):
             try:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                
                # Find header
                header_row_idx = None
                rows_buffer = []
                for i in range(20):
                    try:
                        row = next(rows_iter)
                        rows_buffer.append(row)
                        row_str = " ".join([str(x) for x in row if x is not None])
                        if "Fecha" in row_str and "Tipo" in row_str:
                            header_row_idx = i
                            break
                    except StopIteration:
                        break
                
                if header_row_idx is not None:
                     header = rows_buffer[header_row_idx]
                     col_idx = -1
                     for c_idx, val in enumerate(header):
                         val_str = str(val).lower() if val else ""
                         if "fecha" in val_str: 
                             col_idx = c_idx
                             break
                     
                     if col_idx != -1:
                         # Get next 5 dates
                         for _ in range(5):
                             try:
                                 row = next(rows_iter)
                                 val = row[col_idx]
                                 date_samples.append(f"Sheet '{sheet_name}': {val} (Type: {type(val)})")
                             except StopIteration:
                                 break
             except Exception:
                 pass
        
        wb.close()
        print("\nCaptured Date Samples:")
        for s in date_samples:
            print(s)

    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    debug_dates()
