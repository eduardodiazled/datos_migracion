import openpyxl
import json
import glob
import os
from datetime import datetime

EXCEL_FILES = glob.glob("DATOS/**/*.xlsx", recursive=True)
OUTPUT_FILE = "services_to_restore.json"

restoration_data = []

print(f"Scanning {len(EXCEL_FILES)} Excel files...")

for file in EXCEL_FILES:
    # Processing only recent years to save time/focus
    if "2023" not in file and "2024" not in file and "2025" not in file:
        continue
        
    print(f"Processing {file}...")
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            
            # Buffer first 20 rows to find header
            header_row_idx = None
            rows_buffer = []
            
            for i in range(20):
                try:
                    row = next(rows_iter)
                    rows_buffer.append(row)
                    row_str = " ".join([str(x).lower() for x in row if x is not None])
                    print(f"    [ROW {i}] {row_str}") 
                    if "fecha" in row_str and ("cliente" in row_str or "contacto" in row_str):
                        header_row_idx = i
                        print(f"    Found header at row {i}")
                        break
                except StopIteration:
                    break
            
            if header_row_idx is None:
                continue
                
            # Map columns
            header = rows_buffer[header_row_idx]
            col_map = {}
            for idx, val in enumerate(header):
                val_str = str(val).lower() if val else ""
                if "fecha" in val_str: col_map['date'] = idx
                elif "cliente" in val_str or "contacto" in val_str: col_map['client'] = idx
                elif "monto" in val_str or "valor" in val_str or "total" in val_str or "precio" in val_str: col_map['amount'] = idx
                elif "descripci" in val_str or "servicio" in val_str or "producto" in val_str: col_map['service'] = idx
            
            if 'client' not in col_map or 'amount' not in col_map or 'service' not in col_map:
                print(f"Skipping Sheet {sheet_name}: Missing cols in {col_map}")
                continue

            # Process Data
            data_rows = rows_buffer[header_row_idx+1:]
            
            # Combine buffered rows and remaining iterator
            def row_generator():
                for r in data_rows: yield r
                for r in rows_iter: yield r
                
            for row in row_generator():
                try:
                    # Extract values
                    if len(row) <= max(col_map.values()): continue
                    
                    raw_date = row[col_map['date']]
                    raw_client = row[col_map['client']]
                    raw_amount = row[col_map['amount']]
                    raw_service = row[col_map['service']]
                    
                    # Normalize Date
                    date_val = None
                    if isinstance(raw_date, datetime):
                        date_val = raw_date
                    elif isinstance(raw_date, str):
                        try: date_val = datetime.strptime(raw_date, "%d/%m/%Y")
                        except: pass
                    
                    # Normalize Client (Phone)
                    client_val = None
                    if raw_client:
                        s = str(raw_client).replace('.0', '').strip()
                        if len(s) >= 7: client_val = s
                        
                    # Normalize Amount
                    amount_val = 0
                    if raw_amount:
                        try: amount_val = float(raw_amount)
                        except: pass
                        
                    # Normalize Service
                    service_val = None
                    if raw_service:
                        s = str(raw_service).strip()
                        if s and s.lower() not in ['nan', 'none', '']:
                            service_val = s
                            
                    if date_val and client_val and amount_val > 0 and service_val:
                         restoration_data.append({
                            'phone': client_val,
                            'date': date_val.strftime('%Y-%m-%d'),
                            'amount': amount_val,
                            'service': service_val
                        })
                        
                except Exception as e:
                    continue

    except Exception as e:
        print(f"Error extracting {file}: {e}")

print(f"Found {len(restoration_data)} records to restore.")
with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    json.dump(restoration_data, f, indent=2, ensure_ascii=False)
